import os.path

from xls2xlsx import XLS2XLSX
from openpyxl import load_workbook
from typing import List, Dict, Any
import re
from parser.utils.json_utils import json_converter, save_json, JSONConverter, JSONFileManager


class ScheduleParser:
    def __init__(self, document_path: str) -> None:
        self._document_path = document_path
        if self._document_path.endswith(".xls"):
            self._convert_xls_to_xlsx()
        self._open_worksheet()
        self.converter = JSONConverter
        self.json_manager = JSONFileManager

    def _convert_xls_to_xlsx(self) -> None:
        xls2xlsx = XLS2XLSX(self._document_path)
        self._document_path = f"{os.path.splitext(self._document_path)[0]}.xlsx"
        xls2xlsx.to_xlsx(self._document_path)

    def _open_worksheet(self) -> None:
        self._workbook = load_workbook(self._document_path)
        self._worksheet = self._workbook.worksheets[0]

    def _get_group_row_index(self) -> int:
        index = 0
        for row in self._worksheet.iter_rows():
            index += 1
            for cell in row:
                if cell.value is not None:
                    cell_value = str(cell.value)
                    if re.match(r'^[А-Яа-я]{1,3} - \d{2}', cell_value) is not None:
                        return index

    def get_groups(self) -> List:
        group_row_index = self._get_group_row_index()
        groups: List = []
        for row in self._worksheet.iter_rows(min_row=group_row_index, max_row=group_row_index):
            for cell in row:
                cell_value: str = str(cell.value)
                if cell_value.count('-') >= 1:
                    if cell_value.count('-') > 1:
                        if cell_value.count(',') == 0:
                            cell_value = re.sub('  +', ',', cell_value)
                        cell_value = cell_value.replace(' ', '')
                        cell_value = cell_value.replace(',', ", ")
                    else:
                        cell_value = cell_value.replace(' ', '')
                    groups.append(cell_value)
        print(groups)
        return groups

    def _get_group_coordinates(self, group_row_index: int) -> List:
        study_group_coordinates: List = []
        for row in self._worksheet.iter_rows(min_row=group_row_index, max_row=group_row_index):
            for cell in row:
                if cell.value is None or cell.value.count('-') == 0:
                    continue
                else:
                    study_group_coordinates.append(cell.column)
        return study_group_coordinates

    def _get_lessons_row_index(self, group_row_index) -> int:
        group_columns = self._get_group_coordinates(self._get_group_row_index())
        next_cell = self._worksheet.cell(group_row_index + 1, group_columns[0])
        if next_cell.value is None:
            return group_row_index + 2
        else:
            return group_row_index + 1

    def _end_of_sheet_index(self) -> int:
        index = self._get_lessons_row_index(self._get_group_row_index())
        for row in self._worksheet.iter_rows(min_row=index + 1):
            flag = False
            for cell in row:
                if cell.value is not None:
                    flag = True
            if flag is True:
                index += 1
            else:
                return index - 1

    def _read_sheet(self) -> Dict:
        lesson_time: str | None = None
        day: str | None = None
        lessons: Dict = {}
        group_row_index: int = self._get_group_row_index()
        group_column_coordinates: List = self._get_group_coordinates(group_row_index)

        study_days: List = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
        lesson_numbers: List = ['1', '3', '5', '7', '9', "11"]

        first_row_of_lessons: int = self._get_lessons_row_index(group_row_index)
        end_of_sheet: int = self._end_of_sheet_index()
        for row in self._worksheet.iter_rows(min_row=first_row_of_lessons, max_row=end_of_sheet):
            for cell in row:
                cell_value: Any = cell.value
                if cell_value is None:
                    if cell.column in group_column_coordinates:
                        cell_value = "Нет пары"
                    else:
                        continue
                else:
                    cell_value = str(cell.value)
                if cell_value in lesson_numbers:
                    lesson_time = cell_value + f'-{int(cell_value) + 1} урок'
                    if lesson_time not in lessons[day]:
                        lessons[day][lesson_time] = []
                elif any(day in cell_value.strip().lower() for day in study_days):
                    day = self._format_day(cell_value.strip().lower())
                    if day not in lessons:
                        lessons[day] = {}
                else:
                    if len(cell_value) > 1:
                        subject: str = re.sub(" +", " ", cell_value.strip())
                        lessons[day][lesson_time].append(subject)
        print(lessons)
        return lessons

    def _get_schedule_for_each_group(self) -> Dict:
        groups: List = self.get_groups()
        raw_schedule: Dict = self._read_sheet()
        schedule: Dict = {}
        for group_index in range(len(groups)):
            lesson_index: int = group_index
            schedule[groups[group_index]] = {}
            for day in raw_schedule:
                schedule[groups[group_index]][day] = {}
                for time in raw_schedule[day]:
                    schedule[groups[group_index]][day][time] = raw_schedule[day][time][lesson_index]
        print(schedule)
        return schedule

    @staticmethod
    def _format_day(day: str) -> str:
        formatted_day = day.split(',')[0]
        return formatted_day

    def get_json_schedule(self, file_name: str) -> None:
        schedule = self._get_schedule_for_each_group()
        json_schedule = self.converter.convert_dict_to_json(schedule)
        self.json_manager.save_file(json_file=json_schedule, file_name=file_name)

    def get_json_groups(self, file_name: str) -> None:
        groups = self.get_groups()
        json_groups = self.converter.convert_list_to_json(groups)
        self.json_manager.save_file(json_file=json_groups, file_name=file_name)
